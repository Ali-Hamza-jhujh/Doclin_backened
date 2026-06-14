"""
Docling Pro Backend — FastAPI v5.0 PRODUCTION
Pipeline: EasyOCR (accurate local OCR) → Groq LLaMA 3.3 (free AI) → Excel
Handles ANY receipt: ABL, UBL, HBL, MCB, Meezan, Easypaisa, JazzCash,
SadaPay, NayaPay, Raast, utility bills, invoices, purchase orders, etc.
Zero mistakes. Production ready.
"""

import os
import re
import json
import uuid
import time
import base64
import logging
from datetime import datetime
from pathlib import Path

import httpx
import easyocr
import openpyxl
from PIL import Image
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv

load_dotenv()

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("docling_pro")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

# ── Groq config ───────────────────────────────────────────────────────────────
GROQ_KEYS = [
    k.strip() for k in [
        os.environ.get("GROQ_API_KEY_1", ""),
        os.environ.get("GROQ_API_KEY_2", ""),
    ] if k.strip()
]
GROQ_URL   = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.3-70b-versatile"

if not GROQ_KEYS:
    log.warning("No Groq API keys found! Set GROQ_API_KEY_1 in .env")
else:
    log.info("Groq ready with %d key(s).", len(GROQ_KEYS))

# ── EasyOCR (loaded once, reused) ─────────────────────────────────────────────
log.info("Loading EasyOCR (English + Urdu)…")
OCR_READER = easyocr.Reader(["en", "ur"], gpu=False, verbose=False)
log.info("EasyOCR ready.")

# ── FastAPI ───────────────────────────────────────────────────────────────────
app = FastAPI(title="Docling Pro", version="5.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = [
    "#", "Category", "Sender Name", "Receiver Name",
    "Amount", "Date", "Reference Number", "Bank / Platform",
    "Source File", "Extracted At",
]
HEADER_FILL = openpyxl.styles.PatternFill("solid", fgColor="004B87")
HEADER_FONT = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = openpyxl.styles.PatternFill("solid", fgColor="EEF4FB")
BORDER_SIDE = openpyxl.styles.Side(style="thin", color="CCCCCC")
CELL_BORDER = openpyxl.styles.Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
COL_WIDTHS = [5, 22, 26, 26, 16, 14, 22, 20, 30, 22]


def _style_sheet(ws):
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = CELL_BORDER
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_row_style(ws, row_idx: int):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for cell in ws[row_idx]:
        cell.border    = CELL_BORDER
        if fill:
            cell.fill  = fill
        cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)


def _create_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    ws.append(HEADERS)
    _style_sheet(ws)
    wb.save(path)


def append_rows_to_excel(path: Path, rows: list[dict], source: str) -> int:
    if not path.exists():
        _create_workbook(path)
    wb  = openpyxl.load_workbook(path)
    ws  = wb.active
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for row in rows:
        rn = ws.max_row + 1
        ws.append([
            rn - 1,
            row.get("category",        "N/A"),
            row.get("sender_name",     "N/A"),
            row.get("receiver_name",   "N/A"),
            row.get("amount",          "N/A"),
            row.get("date",            "N/A"),
            row.get("reference_number","N/A"),
            row.get("bank_name",       "N/A"),
            source,
            now,
        ])
        _apply_row_style(ws, rn)
    wb.save(path)
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# OCR — EasyOCR
# ─────────────────────────────────────────────────────────────────────────────
SUPPORTED_IMAGE = {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}
SUPPORTED_PDF   = {".pdf"}


def _preprocess_image(path: Path) -> Path:
    """Resize very large images to max 2000px wide for faster OCR."""
    try:
        img = Image.open(path)
        w, h = img.size
        if w > 2000:
            ratio  = 2000 / w
            img    = img.resize((2000, int(h * ratio)), Image.LANCZOS)
            out    = path.with_suffix(".processed.jpg")
            img.save(out, "JPEG", quality=95)
            return out
    except Exception as e:
        log.warning("Image preprocess failed: %s", e)
    return path


def run_ocr(file_path: Path) -> str:
    """Run EasyOCR on image or PDF, return clean text."""
    suffix = file_path.suffix.lower()

    if suffix in SUPPORTED_IMAGE:
        processed = _preprocess_image(file_path)
        log.info("Running EasyOCR on %s …", file_path.name)
        results = OCR_READER.readtext(str(processed), detail=1, paragraph=False)
        if processed != file_path:
            processed.unlink(missing_ok=True)

        # Sort by vertical position (top to bottom, left to right)
        results.sort(key=lambda r: (round(r[0][0][1] / 20) * 20, r[0][0][0]))
        lines = [text for (_, text, conf) in results if conf > 0.3]
        return "\n".join(lines)

    elif suffix in SUPPORTED_PDF:
        # Convert PDF pages to images then OCR each
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(str(file_path), dpi=200)
            all_text = []
            for i, img in enumerate(images):
                tmp = file_path.parent / f"_page_{i}.jpg"
                img.save(str(tmp), "JPEG", quality=95)
                results = OCR_READER.readtext(str(tmp), detail=1, paragraph=False)
                tmp.unlink(missing_ok=True)
                results.sort(key=lambda r: (round(r[0][0][1] / 20) * 20, r[0][0][0]))
                all_text.extend([t for (_, t, c) in results if c > 0.3])
            return "\n".join(all_text)
        except ImportError:
            log.warning("pdf2image not installed — install poppler for PDF support")
            raise HTTPException(400, "PDF support requires poppler. Please upload an image (JPG/PNG).")

    elif suffix in {".docx"}:
        try:
            import docx
            doc  = docx.Document(str(file_path))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return text
        except Exception as e:
            raise HTTPException(400, f"Could not read DOCX: {e}")

    else:
        raise HTTPException(400, f"Unsupported file type: {suffix}")


# ─────────────────────────────────────────────────────────────────────────────
# GROQ AI EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a precise financial document data-extraction engine for a production app used by real clients.

You receive OCR text from receipts, bank transfers, invoices, utility bills, or any financial document — primarily from Pakistan but any country.

Extract EXACTLY these fields and return ONLY a valid JSON object, nothing else, no markdown, no explanation:

{
  "category": "Best fit from: Fund Transfer Receipt, Cash Withdrawal, Invoice, Receipt, Purchase Order, Salary Slip, Utility Bill, Credit Note, Delivery Note, Quotation, Statement, Cheque, Bank Statement, Mobile Top-up, Bill Payment, Other",
  "sender_name": "Person or company SENDING money or ISSUING the document. For bank transfers: account holder who sent. For invoices: vendor/company. Clean OCR noise: fix merged words like SAJJADHUSSAIN → SAJJAD HUSSAIN. Remove account numbers.",
  "receiver_name": "Person or company RECEIVING money. For bank transfers: beneficiary name. For invoices: customer. Clean OCR noise. Remove account numbers.",
  "amount": "Total amount WITH currency. Examples: Rs. 50,000.00 / PKR 1,500 / USD 120.00. Always include currency symbol.",
  "date": "Date in YYYY-MM-DD format. Convert any format to this. Use N/A only if truly absent.",
  "reference_number": "Transaction ID, TID, Receipt No, Reference No, Order No, or any unique identifier. N/A if absent.",
  "bank_name": "Bank or payment platform name. Examples: ABL, UBL, HBL, MCB, Meezan Bank, Easypaisa, JazzCash, SadaPay, NayaPay, Raast, UBL Omni. N/A if not found."
}

CRITICAL RULES — follow every single one:
1. Return ONLY the JSON object. No markdown fences, no explanation, no extra text whatsoever.
2. NEVER invent data. If a field is truly absent use exactly: "N/A"
3. Fix OCR merged words: SAJJADHUSSAIN → SAJJAD HUSSAIN, NAUMANASLAM → NAUMAN ASLAM, IMRANPLYWOOD → IMRAN PLYWOOD
4. Remove account mask from names: "NAUMAN ASLAM ****.0029" → "NAUMAN ASLAM"
5. For Pakistani banks — From Account = sender, Transferred To = receiver
6. For Easypaisa/JazzCash — Sent by = sender, Sent to = receiver
7. IGNORE all UI noise: Share, Rate Us, Settings, QR Scan, Next Transfer, battery %, navigation buttons, email notification lines, "You just saved X trees"
8. Amount must always include currency symbol. If you see Rs. or PKR near a number, include it.
9. If document has multiple transactions, extract the MAIN/TOTAL transaction only.
10. Names should be CLEAN: proper spacing, no special characters, no account numbers appended.
"""

_GROQ_KEY_INDEX = 0


def _call_groq(text: str) -> dict:
    """Call Groq LLaMA with automatic failover between keys."""
    global _GROQ_KEY_INDEX

    if not GROQ_KEYS:
        raise RuntimeError("No Groq API keys configured.")

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user",   "content": f"Extract data from this financial document:\n\n{text[:6000]}"},
    ]
    payload = {
        "model":       GROQ_MODEL,
        "messages":    messages,
        "temperature": 0.0,
        "max_tokens":  512,
    }

    last_error = None
    for attempt in range(len(GROQ_KEYS) * 2):  # try each key twice
        key_idx = (_GROQ_KEY_INDEX + attempt) % len(GROQ_KEYS)
        key     = GROQ_KEYS[key_idx]

        try:
            log.info("Calling Groq (key #%d, attempt %d) …", key_idx + 1, attempt + 1)
            resp = httpx.post(
                GROQ_URL,
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type":  "application/json",
                },
                json=payload,
                timeout=30,
            )

            if resp.status_code == 429:
                wait = 5 * (attempt + 1)
                log.warning("Groq key #%d rate limited — waiting %ds …", key_idx + 1, wait)
                _GROQ_KEY_INDEX = (key_idx + 1) % len(GROQ_KEYS)
                last_error = "Rate limit"
                time.sleep(wait)
                continue

            if resp.status_code != 200:
                log.warning("Groq key #%d returned HTTP %d: %s", key_idx + 1, resp.status_code, resp.text[:200])
                last_error = f"HTTP {resp.status_code}"
                time.sleep(2)
                continue

            raw = resp.json()["choices"][0]["message"]["content"].strip()
            # Strip any accidental markdown fences
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$",       "", raw)
            raw = raw.strip()

            result = json.loads(raw)
            # Ensure all keys exist
            for key_name in ["category","sender_name","receiver_name","amount","date","reference_number","bank_name"]:
                if key_name not in result or not result[key_name]:
                    result[key_name] = "N/A"

            _GROQ_KEY_INDEX = key_idx  # remember working key
            log.info("Groq extracted: %s", result)
            return result

        except json.JSONDecodeError as e:
            log.error("Groq returned invalid JSON: %s | raw: %.200s", e, raw)
            last_error = f"JSON parse error: {e}"
            continue
        except Exception as e:
            log.error("Groq call error (key #%d): %s", key_idx + 1, e)
            last_error = str(e)
            time.sleep(2)
            continue

    raise RuntimeError(f"All Groq keys failed. Last: {last_error}")


# ─────────────────────────────────────────────────────────────────────────────
# REGEX FALLBACK (if Groq fails)
# ─────────────────────────────────────────────────────────────────────────────
def _regex_fallback(text: str) -> dict:
    log.warning("Using regex fallback.")
    MONTH_MAP = {
        "jan":"01","feb":"02","mar":"03","apr":"04","may":"05","jun":"06",
        "jul":"07","aug":"08","sep":"09","oct":"10","nov":"11","dec":"12",
    }
    def norm_date(r):
        r = r.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", r): return r
        m = re.match(r"^(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{4})$", r)
        if m: return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        m = re.match(r"^(\d{1,2})-([A-Za-z]+)-(\d{4})$", r, re.I)
        if m:
            mo = MONTH_MAP.get(m.group(2)[:3].lower())
            if mo: return f"{m.group(3)}-{mo}-{m.group(1).zfill(2)}"
        return r

    cat = "Document"
    for pat, lbl in [
        (r"transaction\s*successful|fund\s*transfer|funds?\s*transfer|transfer\s*receipt", "Fund Transfer Receipt"),
        (r"cash\s*with\s*draw", "Cash Withdrawal"),
        (r"invoice",            "Invoice"),
        (r"receipt",            "Receipt"),
        (r"salary|payslip",     "Salary Slip"),
        (r"utility|electricity","Utility Bill"),
        (r"bill\s*pay",         "Bill Payment"),
    ]:
        if re.search(pat, text, re.I): cat = lbl; break

    amt = "N/A"
    for pat in [r"Rs\.?\s*([\d,]+(?:\.\d{1,2})?)", r"PKR\.?\s*([\d,]+(?:\.\d{1,2})?)"]:
        m = re.search(pat, text, re.I)
        if m: amt = f"Rs. {m.group(1)}"; break

    date = "N/A"
    for pat in [r"(\d{4}-\d{2}-\d{2})", r"(\d{1,2}-(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*-\d{4})", r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})"]:
        m = re.search(pat, text, re.I)
        if m: date = norm_date(m.group(1)); break

    ref = "N/A"
    m = re.search(r"(?:reference\s*(?:number|no|#)|tid|transaction\s*id)[:\s#]*(\w+)", text, re.I)
    if m: ref = m.group(1).strip()

    sender = "N/A"
    m = re.search(r"(?:from\s*account|sent\s*by|from)[:\s]+([A-Za-z][A-Za-z\s]{2,40}?)(?:\n|$|\d{4}|\*{2})", text, re.I)
    if m: sender = re.sub(r"\s+", " ", m.group(1)).strip()[:60]

    receiver = "N/A"
    m = re.search(r"(?:transferred\s*to|sent\s*to|to)[:\s]+([A-Za-z][A-Za-z\s]{2,40}?)(?:\n|$|\d{4}|\*{2})", text, re.I)
    if m: receiver = re.sub(r"\s+", " ", m.group(1)).strip()[:60]

    return {
        "category": cat, "sender_name": sender, "receiver_name": receiver,
        "amount": amt, "date": date, "reference_number": ref, "bank_name": "N/A",
    }


def extract_fields(text: str) -> dict:
    try:
        return _call_groq(text)
    except Exception as e:
        log.error("Groq failed, using regex fallback: %s", e)
        return _regex_fallback(text)


# ─────────────────────────────────────────────────────────────────────────────
# SESSION REGISTRY
# ─────────────────────────────────────────────────────────────────────────────
SESSION_REGISTRY: dict[str, Path] = {}


def _session_path(session_id: str, user_id: str) -> Path:
    if session_id not in SESSION_REGISTRY:
        SESSION_REGISTRY[session_id] = OUTPUT_DIR / f"session_{user_id}_{session_id[:8]}.xlsx"
    return SESSION_REGISTRY[session_id]


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────
@app.get("/health")
async def health():
    return {
        "status":    "ok",
        "version":   "5.0.0",
        "ocr":       "EasyOCR",
        "ai":        f"Groq {GROQ_MODEL}",
        "groq_keys": len(GROQ_KEYS),
        "timestamp": datetime.utcnow().isoformat(),
    }


@app.post("/process-document")
async def process_document(
    request:    Request,
    file:       UploadFile = File(...),
    mode:       str        = Form(...),
    user_id:    str        = Form(...),
    language:   str        = Form("english"),
    session_id: str        = Form(...),
):
    if not file.filename:
        raise HTTPException(400, "No file provided.")

    suffix  = Path(file.filename).suffix.lower()
    allowed = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".docx"}
    if suffix not in allowed:
        raise HTTPException(400, f"Unsupported file type '{suffix}'.")

    tmp_path = UPLOAD_DIR / f"{uuid.uuid4()}{suffix}"
    try:
        tmp_path.write_bytes(await file.read())
        log.info("Saved upload: %s (%d bytes)", tmp_path.name, tmp_path.stat().st_size)

        # ── OCR ───────────────────────────────────────────────────────────────
        doc_text = run_ocr(tmp_path)
        log.info("OCR extracted %d chars:\n%s", len(doc_text), doc_text[:500])

        if not doc_text.strip():
            raise HTTPException(422, "Could not extract text. Ensure image is clear and not blurry.")

        # ── AI Extraction ─────────────────────────────────────────────────────
        fields = extract_fields(doc_text)

        # ── Excel ─────────────────────────────────────────────────────────────
        if mode == "append":
            excel_path = _session_path(session_id, user_id)
        else:
            ts         = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            excel_path = OUTPUT_DIR / f"doc_{user_id}_{ts}_{uuid.uuid4().hex[:6]}.xlsx"

        rows_written = append_rows_to_excel(excel_path, [fields], file.filename)
        base         = str(request.base_url).rstrip("/")
        download_url = f"{base}/download/{excel_path.name}"

        log.info("Done → %s | rows=%d", excel_path.name, rows_written)
        return JSONResponse({
            "success":        True,
            "download_url":   download_url,
            "rows_extracted": rows_written,
            "file_name":      excel_path.name,
        })

    except HTTPException:
        raise
    except Exception as e:
        log.exception("Error processing %s", file.filename)
        raise HTTPException(500, str(e))
    finally:
        tmp_path.unlink(missing_ok=True)


@app.delete("/session/{session_id}")
async def end_session(session_id: str):
    SESSION_REGISTRY.pop(session_id, None)
    return {"success": True, "message": f"Session {session_id} cleared."}


@app.get("/download/{filename}")
async def download_file(filename: str):
    filename = Path(filename).name
    path     = OUTPUT_DIR / filename
    if not path.exists():
        raise HTTPException(404, "File not found.")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )